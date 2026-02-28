"""
Read all raw Apollo CSV exports, parse (handle quoted newlines), deduplicate,
and write one cleaned Excel file (.xlsx).
"""
import csv
from pathlib import Path

from openpyxl import Workbook

SEGMENT = "lead_gen_appointment_setting"
RAW_FILES = [
    "apollo-accounts-export.csv",
    "apollo-accounts-export (1).csv",
    "apollo-accounts-export (2).csv",
    "apollo-accounts-export (3).csv",
    "apollo-accounts-export (4).csv",
    "apollo-accounts-export (5).csv",
    "apollo-accounts-export (6).csv",
]
OUTPUT_XLSX = "apollo-accounts-export-cleaned.xlsx"


def parse_one_csv(in_path):
    """Read one Apollo CSV; return (header, list of data rows). Handles quoted newlines."""
    with open(in_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return None, []
    header = list(rows[0])
    if "segment" not in header:
        header.append("segment")
    data = []
    for i in range(1, len(rows)):
        row = rows[i]
        while len(row) < len(header) - 1:
            row.append("")
        if len(row) == len(header) - 1:
            row.append(SEGMENT)
        elif len(row) > len(header):
            row = row[: len(header) - 1] + [SEGMENT]
        else:
            row = row + [SEGMENT]
        row = [str(c).replace("\r\n", " ").replace("\n", " ").replace("\r", " ").strip() for c in row]
        data.append(row[: len(header)])
    return header, data


def main():
    base = Path(__file__).resolve().parent.parent  # repo root
    all_header = None
    all_rows = []
    seen_ids = set()

    for name in RAW_FILES:
        path = base / name
        if not path.exists():
            print(f"Skip (not found): {path}")
            continue
        header, rows = parse_one_csv(path)
        if header is None:
            continue
        if all_header is None:
            all_header = header
        # Deduplicate by Apollo Account Id (keep first occurrence)
        try:
            idx = all_header.index("Apollo Account Id")
        except ValueError:
            idx = all_header.index("Company Name")  # fallback
        for row in rows:
            if len(row) <= idx:
                all_rows.append(row)
                continue
            key = (row[idx] or "").strip()
            if key and key in seen_ids:
                continue
            if key:
                seen_ids.add(key)
            all_rows.append(row)

    if not all_header or not all_rows:
        print("No data to write.")
        return

    # Add hubspot / salesforce columns: True if word appears anywhere in record (case insensitive)
    all_header = all_header + ["hubspot", "salesforce"]
    for row in all_rows:
        record_text = " ".join(str(c) for c in row).lower()
        row.append("hubspot" in record_text)
        row.append("salesforce" in record_text)

    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"
    for col, name in enumerate(all_header, start=1):
        ws.cell(row=1, column=col, value=name)
    for r, row in enumerate(all_rows, start=2):
        for c, val in enumerate(row, start=1):
            if c <= len(all_header):
                ws.cell(row=r, column=c, value=val)

    out_path = base / OUTPUT_XLSX
    wb.save(out_path)
    print(f"Wrote {len(all_rows)} companies to {out_path} (Excel format, deduplicated).")


if __name__ == "__main__":
    main()
