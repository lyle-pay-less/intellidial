"""
Phase 2: Use VAPI to call gynecologists and gather information
Saves results to Excel and tracks called numbers
"""
import requests
import json
import time
import os
from openpyxl import Workbook, load_workbook
from config import VAPI_API_KEY, VAPI_PHONE_NUMBER_ID, GEMINI_API_KEY

VAPI_BASE_URL = "https://api.vapi.ai"

# File to track which numbers we've already called
CALLED_NUMBERS_FILE = "called_numbers.json"

# The assistant prompt - this is what the AI will say on the call
ASSISTANT_CONFIG = {
    "name": "Gyni Finder Assistant",
    "model": {
        "provider": "openai",
        "model": "gpt-4o-mini",
        "messages": [
            {
                "role": "system",
                "content": """You are calling medical practices to find a gynecologist appointment.

YOUR NAME: If they ask your name, say "Zoey". Only say your name if asked.

PATIENT INFO (only share if they ask for details):
- 24-year-old woman
- Had an abortion pill on 15 January
- Went back to the doctor on the 20th because she wasn't feeling well
- They found an infection and gave her medication
- Completed the antibiotics
- Still has some nausea

YOUR GOAL - gather this information:
1. Ask if they have a gynecologist available for new patients
2. Ask if they offer ultrasound services  
3. Ask the consultation price
4. Ask about the EARLIEST available appointment (specific date and time)

CRITICAL - IVR/AUTOMATED SYSTEMS:
- If you hear "dial 1", "press 1", "for medical appointments", etc - SAY "1" or press the number
- If you hear menu options, select the one for medical/appointments (usually 1)
- NEVER end the call when you hear an automated system - navigate through it
- Wait until you reach a real human receptionist

CRITICAL - DO NOT END EARLY:
- You MUST ask all 4 questions before ending: gyni available? ultrasound? price? earliest appointment?
- NEVER say goodbye until you have answers to these questions
- If you haven't asked about price yet, DO NOT END THE CALL
- Only end after getting all information OR the person explicitly says they cannot help with gynecology

CONVERSATION FLOW:
1. Greet and wait
2. Ask: "Do you have a gynecologist available for new patients?"
3. Ask: "Do you offer ultrasound services?"
4. Ask: "What is the consultation fee?"
5. Ask: "When is the earliest available appointment?"
6. ONLY THEN say thank you and end
"""
            }
        ]
    },
    "voice": {
        "provider": "11labs",
        "voiceId": "21m00Tcm4TlvDq8ikWAM"  # Rachel - professional female voice
    },
    "firstMessage": "Hi, good day!",
    "endCallFunctionEnabled": False,
    "transcriber": {
        "provider": "deepgram",
        "model": "nova-2",
        "language": "en"
    },
    "maxDurationSeconds": 180,  # 3 minute max per call
    "silenceTimeoutSeconds": 30,
    "responseDelaySeconds": 0.5,
}


def get_headers():
    return {
        "Authorization": f"Bearer {VAPI_API_KEY}",
        "Content-Type": "application/json"
    }


def download_recording(recording_url: str, practice_name: str, call_id: str) -> str:
    """Download the call recording and save locally"""
    if not recording_url:
        return ""
    
    # Create recordings folder if it doesn't exist
    if not os.path.exists(RECORDINGS_FOLDER):
        os.makedirs(RECORDINGS_FOLDER)
    
    # Clean practice name for filename
    safe_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in practice_name)
    safe_name = safe_name[:50]  # Limit length
    filename = f"{safe_name}_{call_id[:8]}.mp3"
    filepath = os.path.join(RECORDINGS_FOLDER, filename)
    
    try:
        response = requests.get(recording_url, stream=True)
        if response.status_code == 200:
            with open(filepath, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            print(f"  ✓ Recording saved: {filepath}")
            return filepath
        else:
            print(f"  ✗ Failed to download recording: {response.status_code}")
            return ""
    except Exception as e:
        print(f"  ✗ Recording download error: {e}")
        return ""


def load_called_numbers() -> set:
    """Load the set of phone numbers we've already called"""
    try:
        with open(CALLED_NUMBERS_FILE, "r") as f:
            return set(json.load(f))
    except FileNotFoundError:
        return set()


def save_called_number(phone: str):
    """Add a phone number to the called list"""
    called = load_called_numbers()
    called.add(phone)
    with open(CALLED_NUMBERS_FILE, "w") as f:
        json.dump(list(called), f)


EXCEL_FILE = "gyni_results.xlsx"
RECORDINGS_FOLDER = "recordings"

def setup_excel():
    """Setup Excel file with headers"""
    headers = [
        "Practice Name",
        "Phone",
        "Address", 
        "Gyni Available",
        "Ultrasound Available",
        "Consultation Price",
        "Earliest Availability",
        "Call Status",
        "Call Duration (s)",
        "Full Transcript",
        "Recording URL",
        "Local Recording File",
        "Called At"
    ]
    
    if os.path.exists(EXCEL_FILE):
        # Load existing workbook
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        print(f"Loaded existing Excel file: {EXCEL_FILE}")
    else:
        # Create new workbook with headers
        wb = Workbook()
        ws = wb.active
        ws.title = "Call Results"
        ws.append(headers)
        wb.save(EXCEL_FILE)
        print(f"Created new Excel file: {EXCEL_FILE}")
    
    return wb, ws


def append_to_excel(wb, ws, result: dict):
    """Append a result row to Excel"""
    practice = result.get("practice", {})
    
    row = [
        practice.get("name", ""),
        practice.get("phone", ""),
        practice.get("address", ""),
        result.get("has_gyni", "Unknown"),
        result.get("has_ultrasound", "Unknown"),
        result.get("price", "Unknown"),
        result.get("availability", "Unknown"),
        result.get("status", ""),
        result.get("duration_seconds", ""),
        result.get("transcript", ""),
        result.get("recording_url", ""),
        result.get("local_recording", ""),
        time.strftime("%Y-%m-%d %H:%M:%S")
    ]
    
    try:
        ws.append(row)
        wb.save(EXCEL_FILE)
        print(f"  ✓ Added to {EXCEL_FILE}")
    except Exception as e:
        print(f"  ✗ Excel error: {e}")


def create_assistant():
    """Create a VAPI assistant for the calls"""
    url = f"{VAPI_BASE_URL}/assistant"
    response = requests.post(url, headers=get_headers(), json=ASSISTANT_CONFIG)
    
    if response.status_code != 201:
        print(f"Error creating assistant: {response.status_code}")
        print(response.text)
        return None
    
    assistant = response.json()
    print(f"Created assistant: {assistant['id']}")
    return assistant


def make_call(phone_number: str, assistant_id: str, practice_name: str):
    """Make an outbound call to a phone number"""
    url = f"{VAPI_BASE_URL}/call/phone"
    
    # Format phone number for South Africa if needed
    formatted_number = phone_number.strip().replace(" ", "")
    if formatted_number.startswith("0"):
        formatted_number = "+27" + formatted_number[1:]
    elif not formatted_number.startswith("+"):
        formatted_number = "+27" + formatted_number
    
    payload = {
        "phoneNumberId": VAPI_PHONE_NUMBER_ID,
        "assistantId": assistant_id,
        "customer": {
            "number": formatted_number,
            "name": practice_name[:40]  # VAPI limit is 40 chars
        },
        "metadata": {
            "practice_name": practice_name,
            "original_phone": phone_number
        }
    }
    
    response = requests.post(url, headers=get_headers(), json=payload)
    
    if response.status_code != 201:
        print(f"  Error calling {practice_name}: {response.status_code}")
        print(f"  {response.text}")
        return None
    
    call = response.json()
    print(f"  Call initiated: {call['id']} -> {formatted_number}")
    return call


def get_call_status(call_id: str):
    """Get the status of a call"""
    url = f"{VAPI_BASE_URL}/call/{call_id}"
    response = requests.get(url, headers=get_headers())
    
    if response.status_code != 200:
        return None
    
    return response.json()


def wait_for_call_completion(call_id: str, timeout: int = 300):
    """Wait for a call to complete and return the result"""
    start_time = time.time()
    
    while time.time() - start_time < timeout:
        call = get_call_status(call_id)
        
        if not call:
            return None
        
        status = call.get("status")
        
        if status in ["ended", "failed"]:
            return call
        
        time.sleep(5)  # Check every 5 seconds
    
    return None


def analyze_transcript(transcript: str) -> dict:
    """Use Gemini to analyze transcript and extract key info"""
    
    result = {
        "has_gyni": "Unknown",
        "has_ultrasound": "Unknown",
        "price": "Unknown",
        "availability": "Unknown"
    }
    
    if not transcript or not GEMINI_API_KEY:
        return result
    
    try:
        prompt = """Analyze this phone call transcript and extract the following information. 
Return ONLY a JSON object with these exact keys:
- has_gyni: "Yes", "No", or "Unknown"
- has_ultrasound: "Yes", "No", or "Unknown"  
- price: The consultation price mentioned (e.g., "R2800") or "Unknown"
- availability: The earliest available appointment mentioned (e.g., "Wednesday 2pm", "Next Monday") or "Unknown"

Be thorough - look for any mention of prices, fees, costs, availability, appointments, gynecologist/gynaecologist availability.

Transcript:
""" + transcript
        
        response = requests.post(
            f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GEMINI_API_KEY}",
            headers={"Content-Type": "application/json"},
            json={
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {"temperature": 0}
            }
        )
        
        if response.status_code == 200:
            content = response.json()["candidates"][0]["content"]["parts"][0]["text"]
            # Parse the JSON response
            import json
            # Clean up the response (remove markdown code blocks if present)
            content = content.strip()
            if content.startswith("```"):
                content = content.split("```")[1]
                if content.startswith("json"):
                    content = content[4:]
            content = content.strip()
            
            parsed = json.loads(content)
            result["has_gyni"] = parsed.get("has_gyni", "Unknown")
            result["has_ultrasound"] = parsed.get("has_ultrasound", "Unknown")
            result["price"] = parsed.get("price", "Unknown")
            result["availability"] = parsed.get("availability", "Unknown")
            print(f"  AI Analysis: Gyni={result['has_gyni']}, Ultrasound={result['has_ultrasound']}, Price={result['price']}, Avail={result['availability']}")
        else:
            print(f"  Gemini error: {response.status_code} - {response.text}")
    except Exception as e:
        print(f"  AI analysis error: {e}")
    
    return result


def extract_call_results(call: dict) -> dict:
    """Extract relevant information from a completed call"""
    transcript = call.get("transcript", "")
    
    # Analyze transcript for key information
    analysis = analyze_transcript(transcript)
    
    result = {
        "call_id": call.get("id"),
        "status": call.get("status"),
        "duration_seconds": call.get("duration"),
        "end_reason": call.get("endedReason"),
        "transcript": transcript,
        "recording_url": call.get("recordingUrl", ""),
        "cost": call.get("cost"),
        "has_gyni": analysis["has_gyni"],
        "has_ultrasound": analysis["has_ultrasound"],
        "price": analysis["price"],
        "availability": analysis["availability"],
    }
    
    return result


def load_gynecologists(filename: str = "gynecologists.json") -> list:
    """Load the list of gynecologists from Phase 1"""
    try:
        with open(filename, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Error: {filename} not found. Run Phase 1 first.")
        return []


def save_results(results: list, filename: str = "call_results.json"):
    """Save call results to JSON"""
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"\nSaved {len(results)} results to {filename}")


def main():
    if not VAPI_API_KEY:
        print("ERROR: Please set VAPI_API_KEY in your .env file")
        return
    
    if not VAPI_PHONE_NUMBER_ID:
        print("ERROR: Please set VAPI_PHONE_NUMBER_ID in your .env file")
        return
    
    print("=" * 50)
    print("Gyni Finder - Phase 2: VAPI Calling")
    print("=" * 50)
    
    # Setup Excel file
    wb, ws = setup_excel()
    
    # Load gynecologists from Phase 1
    practices = load_gynecologists()
    if not practices:
        return
    
    # Load already-called numbers to avoid duplicates
    called_numbers = load_called_numbers()
    print(f"Already called: {len(called_numbers)} numbers")
    
    # Filter out already-called numbers
    practices_to_call = [p for p in practices if p['phone'] not in called_numbers]
    print(f"Remaining to call: {len(practices_to_call)} practices")
    
    if not practices_to_call:
        print("\n✓ All practices have been called!")
        return
    
    # Create assistant
    print("\nCreating VAPI assistant...")
    assistant = create_assistant()
    if not assistant:
        return
    
    # Limit calls for testing (IMPORTANT: start small!)
    MAX_CALLS = 10  # Batch of 10 calls per run
    practices_batch = practices_to_call[:MAX_CALLS]
    
    print(f"\n📞 Calling {MAX_CALLS} practice(s) this run")
    print("Edit MAX_CALLS in the script to adjust batch size\n")
    
    all_results = []
    
    for i, practice in enumerate(practices_batch, 1):
        print(f"\n[{i}/{len(practices_batch)}] Calling: {practice['name']}")
        print(f"  Phone: {practice['phone']}")
        
        # Make the call
        call = make_call(practice['phone'], assistant['id'], practice['name'])
        
        if not call:
            result = {
                "practice": practice,
                "status": "failed",
                "error": "Failed to initiate call"
            }
            all_results.append(result)
            append_to_excel(wb, ws, result)
            # DON'T mark as called if it failed - so we can retry
            continue
        
        # Wait for call to complete
        print("  Waiting for call to complete...")
        completed_call = wait_for_call_completion(call['id'])
        
        if completed_call:
            result = extract_call_results(completed_call)
            result["practice"] = practice
            
            print(f"  Call ended: {result['end_reason']} ({result['duration_seconds']}s)")
            print(f"  Gyni: {result['has_gyni']} | Ultrasound: {result['has_ultrasound']} | Price: {result['price']}")
            
            # Download the recording
            if result.get("recording_url"):
                local_recording = download_recording(
                    result["recording_url"], 
                    practice['name'], 
                    result['call_id']
                )
                result["local_recording"] = local_recording
            
            all_results.append(result)
            
            # Save to Excel
            append_to_excel(wb, ws, result)
            # Only mark as called if the call actually connected
            save_called_number(practice['phone'])
        else:
            result = {
                "practice": practice,
                "call_id": call['id'],
                "status": "timeout",
                "error": "Call timed out"
            }
            all_results.append(result)
            append_to_excel(wb, ws, result)
            # DON'T mark as called if it timed out - so we can retry
        
        # Small delay between calls
        if i < len(practices_batch):
            print("  Waiting 10s before next call...")
            time.sleep(10)
    
    # Save results to JSON as backup
    save_results(all_results)
    
    print("\n" + "=" * 50)
    print("CALLING COMPLETE")
    print("=" * 50)
    print(f"Calls attempted: {len(all_results)}")
    print(f"Total called so far: {len(load_called_numbers())}")
    print(f"Remaining: {len(practices_to_call) - len(practices_batch)}")
    print("\nResults saved to:")
    print(f"  - {EXCEL_FILE}")
    print(f"  - {RECORDINGS_FOLDER}/ (audio recordings)")
    print("  - call_results.json (backup)")
    print("\nRun the script again to call more practices.")


if __name__ == "__main__":
    main()
