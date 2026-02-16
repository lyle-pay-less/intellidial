#!/usr/bin/env python3
"""Create an Excel lead response tracker for the AutoTrader dealership test."""
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT = SCRIPT_DIR / "lead_response_tracker.xlsx"

LISTINGS = [
    {"car": "VW Amarok 2.0BiTDi", "url": "https://www.autotrader.co.za/car-for-sale/volkswagen/amarok/2.0bitdi/28417706"},
    {"car": "Haval H6 GT Luxury", "url": "https://www.autotrader.co.za/car-for-sale/haval/h6-gt/luxury/28417304"},
    {"car": "Jetour Dashing 1.5T", "url": "https://www.autotrader.co.za/car-for-sale/jetour/dashing/1.5t/28417282"},
    {"car": "Kia Sorento 2.2", "url": "https://www.autotrader.co.za/car-for-sale/kia/sorento/2.2/28415393"},
    {"car": "Hyundai Grand i10 Motion", "url": "https://www.autotrader.co.za/car-for-sale/hyundai/grand-i10/motion/28414153"},
    {"car": "GWM P-Series LTD", "url": "https://www.autotrader.co.za/car-for-sale/gwm/p-series/ltd/28408961"},
    {"car": "Mercedes SLK 350", "url": "https://www.autotrader.co.za/car-for-sale/mercedes-benz/slk/slk350/28413478"},
    {"car": "VW Amarok 3.0TDi", "url": "https://www.autotrader.co.za/car-for-sale/volkswagen/amarok/3.0tdi/28413458"},
    {"car": "Jetour X70 Plus 1.5T", "url": "https://www.autotrader.co.za/car-for-sale/jetour/x70-plus/1.5t/28413186"},
    {"car": "VW T-Cross 1.5TSi", "url": "https://www.autotrader.co.za/car-for-sale/volkswagen/t-cross/1.5tsi/28411127"},
    {"car": "VW Tiguan 2.0TSi", "url": "https://www.autotrader.co.za/car-for-sale/volkswagen/tiguan/2.0tsi/28415602"},
    {"car": "VW Polo GTI", "url": "https://www.autotrader.co.za/car-for-sale/volkswagen/polo/gti/28410031"},
    {"car": "Hyundai Kona Executive", "url": "https://www.autotrader.co.za/car-for-sale/hyundai/kona/executive/28413524"},
    {"car": "Ford Tourneo Custom", "url": "https://www.autotrader.co.za/car-for-sale/ford/tourneo-custom/ambiente/28396761"},
    {"car": "Isuzu D-Max 1.9", "url": "https://www.autotrader.co.za/car-for-sale/isuzu/d-max/1.9/28405017"},
    {"car": "Toyota Fortuner 2.4GD-6", "url": "https://www.autotrader.co.za/car-for-sale/toyota/fortuner/2.4gd-6/28417581"},
    {"car": "Land Rover Evoque D200", "url": "https://www.autotrader.co.za/car-for-sale/land-rover/range-rover-evoque/d200/28411043"},
    {"car": "Isuzu D-Max 250", "url": "https://www.autotrader.co.za/car-for-sale/isuzu/d-max/250/28409776"},
    {"car": "Land Rover Defender 110", "url": "https://www.autotrader.co.za/car-for-sale/land-rover/defender/110/28406113"},
    {"car": "Kia Picanto 1.2", "url": "https://www.autotrader.co.za/car-for-sale/kia/picanto/1.2/28415590"},
]

def main():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Installing openpyxl...")
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Lead Response Test"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title row
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "AutoTrader Lead Response Test - Cape Town Dealers - 11 Feb 2026"
    title_cell.font = Font(bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # Headers
    headers = [
        "#",
        "Car Listing",
        "Dealer Name",
        "AutoTrader URL",
        "Time Enquiry Sent",
        "Got Callback?",
        "Time of Callback",
        "Response Time",
        "Notes",
    ]
    ws.row_dimensions[3].height = 25
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for i, listing in enumerate(LISTINGS, 1):
        row = i + 3
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=listing["car"]).border = thin_border
        ws.cell(row=row, column=3, value="").border = thin_border  # Dealer name - fill in
        ws.cell(row=row, column=4, value=listing["url"]).border = thin_border
        ws.cell(row=row, column=5, value="").border = thin_border  # Time sent
        ws.cell(row=row, column=6, value="").border = thin_border  # Got callback?
        ws.cell(row=row, column=7, value="").border = thin_border  # Time of callback
        ws.cell(row=row, column=8, value="").border = thin_border  # Response time
        ws.cell(row=row, column=9, value="").border = thin_border  # Notes

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 35

    # Summary section
    summary_row = len(LISTINGS) + 5
    ws.merge_cells(f"A{summary_row}:I{summary_row}")
    summary_title = ws.cell(row=summary_row, column=1, value="SUMMARY")
    summary_title.font = Font(bold=True, size=12, color="2F5496")

    summary_items = [
        "Total enquiries sent:",
        "Called back within 1 hour:",
        "Called back within 4 hours:",
        "Called back next day:",
        "Never called back:",
        "Average response time:",
    ]
    for j, item in enumerate(summary_items):
        r = summary_row + 1 + j
        ws.cell(row=r, column=1, value=item).font = Font(bold=True)
        ws.merge_cells(f"A{r}:C{r}")
        if j == 0:
            ws.cell(row=r, column=4, value=20)

    # Pitch lines section
    pitch_row = summary_row + len(summary_items) + 3
    ws.merge_cells(f"A{pitch_row}:I{pitch_row}")
    pitch_title = ws.cell(row=pitch_row, column=1, value="PITCH LINES FROM THIS TEST")
    pitch_title.font = Font(bold=True, size=12, color="2F5496")

    pitches = [
        '"I enquired about your [car] at [time]. Your team called me back [X hours] later. By then I\'d already spoken to 3 other dealers."',
        '"Out of 20 dealers I tested, only [X] called back within an hour. The other [Y] lost the sale."',
        '"What if every lead got a call back in 60 seconds — even at 9pm on a Friday?"',
    ]
    for j, pitch in enumerate(pitches):
        r = pitch_row + 1 + j
        ws.merge_cells(f"A{r}:I{r}")
        ws.cell(row=r, column=1, value=pitch).font = Font(italic=True)

    wb.save(OUTPUT)
    print(f"Created tracker: {OUTPUT}")


if __name__ == "__main__":
    main()
